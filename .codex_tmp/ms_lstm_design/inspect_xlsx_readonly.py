from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def scalar(value: object) -> object:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


for raw_path in sys.argv[1:]:
    path = Path(raw_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    print(json.dumps({"type": "workbook", "path": str(path), "sheets": workbook.sheetnames}, ensure_ascii=False))
    for worksheet in workbook.worksheets:
        rows = []
        for row in worksheet.iter_rows(min_row=1, max_row=min(8, worksheet.max_row), values_only=True):
            rows.append([scalar(value) for value in row[: min(8, worksheet.max_column)]])
        print(
            json.dumps(
                {
                    "type": "sheet_sample",
                    "sheet": worksheet.title,
                    "max_row": worksheet.max_row,
                    "max_column": worksheet.max_column,
                    "rows": rows,
                },
                ensure_ascii=False,
            )
        )
    workbook.close()
