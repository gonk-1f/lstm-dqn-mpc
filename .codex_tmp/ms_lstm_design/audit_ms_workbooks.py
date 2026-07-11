from __future__ import annotations

import json
import math
import sys
from pathlib import Path

import numpy as np
from openpyxl import load_workbook


HISTORY_LEN = 30
PRED_HORIZON = 6


def load_segment(worksheet) -> np.ndarray:
    rows = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        rows.append(row[:5])
    return np.asarray(rows, dtype=np.float64)


for raw_path in sys.argv[1:]:
    path = Path(raw_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    segments = []
    for worksheet in workbook.worksheets[1:]:
        values = load_segment(worksheet)
        time_s = values[:, 0]
        load_kw = values[:, 1]
        sampled = values[::10]
        dt = np.diff(time_s)
        sampled_dt = np.diff(sampled[:, 0])
        segment = {
            "sheet": worksheet.title,
            "rows_1ms": int(len(values)),
            "rows_10ms": int(len(sampled)),
            "windows_30_to_6": max(0, int(len(sampled) - HISTORY_LEN - PRED_HORIZON + 1)),
            "time_start_s": float(time_s[0]),
            "time_end_s": float(time_s[-1]),
            "dt_1ms_min": float(np.min(dt)),
            "dt_1ms_median": float(np.median(dt)),
            "dt_1ms_max": float(np.max(dt)),
            "dt_10ms_min": float(np.min(sampled_dt)),
            "dt_10ms_median": float(np.median(sampled_dt)),
            "dt_10ms_max": float(np.max(sampled_dt)),
            "load_min_kw": float(np.min(load_kw)),
            "load_max_kw": float(np.max(load_kw)),
            "load_mean_kw": float(np.mean(load_kw)),
            "nonfinite_values": int(np.size(values) - np.count_nonzero(np.isfinite(values))),
        }
        segments.append((segment, sampled))
        print(json.dumps({"type": "segment", "workbook": path.name, **segment}, ensure_ascii=False))

    overlap_rows = []
    for left_idx, (left_meta, left) in enumerate(segments):
        left_time = np.rint(left[:, 0] * 1000).astype(np.int64)
        left_map = {int(t): float(v) for t, v in zip(left_time, left[:, 1])}
        for right_meta, right in segments[left_idx + 1 :]:
            right_time = np.rint(right[:, 0] * 1000).astype(np.int64)
            common = np.intersect1d(left_time, right_time, assume_unique=True)
            if len(common) == 0:
                continue
            max_abs_load_delta = max(
                abs(left_map[int(t)] - float(right[np.searchsorted(right_time, t), 1])) for t in common
            )
            overlap_rows.append(
                {
                    "left": left_meta["sheet"],
                    "right": right_meta["sheet"],
                    "overlap_10ms_rows": int(len(common)),
                    "max_abs_load_delta_kw": float(max_abs_load_delta),
                }
            )

    summary = {
        "type": "workbook_summary",
        "workbook": path.name,
        "segments": len(segments),
        "rows_1ms": int(sum(meta["rows_1ms"] for meta, _ in segments)),
        "rows_10ms": int(sum(meta["rows_10ms"] for meta, _ in segments)),
        "windows_30_to_6": int(sum(meta["windows_30_to_6"] for meta, _ in segments)),
        "overlaps": overlap_rows,
    }
    print(json.dumps(summary, ensure_ascii=False))
    workbook.close()
